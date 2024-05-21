mpfrom os.path import join, exists
from os import mkdir
import pandas as pd
import datetime as dt
import openpyxl as oxl
from openpyxl.styles import Border, Side

# Variables
path = "C:\\Users\\correo.ejemplo\\OneDrive - GANA S.A\\NotasTrasladosPY"
template = "Plantilla_PDF_EPM.xlsx"
report_name = "reporte de notas.xls"
onedrive_path = "C:\\Users\\correo.correo\\OneDrive\NotasTrasladosPY\\Onedrive"
products = {'EPM': ('RECAUDOS EPM EN LINEA', 'RECAUDO PAGA A TU MEDIDA', 'REC EPM EN LINEA')}

# Nombre de los reportes xls generados.
filename = 'Consolidado_EPM'
# Nombre de la carpeta donde serán almacenados los reportes generados.
folder_name = 'epm'

global today

# Ruta al reporte de notas (carpeta de OneDrive)
report_path = join(onedrive_path, report_name)

# Ruta al archivo plantilla PDF_EPM
template_path = join(path, "templates", template)

# Declaración de una variable para la fecha de hoy
today = dt.datetime.today().strftime("%d-%m-%Y")

# Crear la carpeta por fecha si no existe
folder = join(path, "reports", today)
if not exists(folder):
    mkdir(folder)

# Crear carpeta para consolidado
folder = join(folder, folder_name)
if not exists(folder):
    mkdir(folder)

# Lectura del reporte de la hoja de reporte de notas.
with open(report_path, mode='rb') as fp:
    df = pd.read_excel(fp, sheet_name='Hoja 1', dtype={'Nro Caso': str})

# Filtrar por columna 'Nro Caso' diferente a 'canal' o 'CANAL'
df_filtered = df[~df['Nro Caso'].str.contains('CANAL|canal', na=False)]

# Crear un Dataframe para productos EPM
epm_products = products.get('EPM')
epm_products = '|'.join(epm_products)
df_filtered = df_filtered[df_filtered['Producto'].str.contains(epm_products, na=False)]
df_filtered.reset_index(drop=True, inplace=True)

# Plantilla de consolidado epm
# Workbook de la plantilla consolidado epm
workbook = oxl.load_workbook(template_path)
# Seleccionar la primera hoja como worksheet
worksheet = workbook.active

cells = {
    "id": ("Id. Nota", 2),
    "centro": ("Oficina", 3),
    "naturaleza": ("Naturaleza", 4),
    "caso": ("Nro Caso", 5),
    "producto": ("Producto", 6),
    "responsable": ("Responsable", 7),
    "obs": ("Observaciones", 8),
    "valor": ("Valor", 9),
    "aliado": ("EPM", 10)
}

# Estilo de las celdas
border_style = Border(
    left=Side(border_style='thin', color='FF000000'),
    right=Side(border_style='thin', color='FF000000'),
    top=Side(border_style='thin', color='FF000000'),
    bottom=Side(border_style='thin', color='FF000000'),
)

# Rango de la plantilla
min_row = 11
min_col = 2
max_col = 10
max_row = min_row

if df_filtered.shape[0] > 1:
    max_row = min_row + df_filtered.shape[0] - 1
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border_style

total = 0
worksheet.cell(1, 6, f'{today}')

for index, row in df_filtered.iterrows():
    for col_name, col in cells.values():
        if col_name == 'EPM':
            worksheet.cell(row=index + min_row, column=col, value=col_name)
        elif col_name == 'Valor':
            total += int(row[col_name])
            worksheet.cell(row=index + min_row, column=col, value=f'${row[col_name]:,}')
        else:
            worksheet.cell(row=index + min_row, column=col, value=row[col_name])

total_cell = worksheet.cell(row=max_row + 1, column=cells.get('valor')[1], value=f'${total:,}')
total_cell.border = border_style
worksheet.cell(4, 6, f'${total:,}')

workbook.save(join(folder, f'{filename}_{today}.xlsx'))
workbook.close()

print(f'{filename}_{today} saved.')
