from os.path import join, exists
from os import mkdir
import pandas as pd
import datetime as dt
import openpyxl as oxl

path = "C:\\Users\\correo.automatizacio\\OneDrive - GANA S.A\\NotasTrasladosPY"
report_name = "reporte de notas.xls"
onedrive_path = "C:\\Users\\correo.automatizacio\\OneDrive - GANA S.A\\NotasTrasladosPY\\Onedrive"
template = "Plantilla_PDF_Nota.xlsx"
templates_path = join(path, "templates")
products = {'OTROS': ('REC LEBON', 'REC DOLCE', 'REC FUNERALES ANTIOQ', 'REC PLACE TO PAY', 'RECAUDO CIA IDIOMAS', 'ZONA VIRTUAL\nREC NATURA ', 'REC JAMAR', 'RECAUDO AGAVAL', 'RECAUDO MEDIA NARANJA', 'REC RINKU CONEKTO', 'RECAUDO FUCSIA FUCSIA', 'RECAUDO VELONET', 'RECAUDO FLYPASS', 'RECAUDO INTERACTUAR', 'REC COOGRANADA', 'REC FED CAFETEROS', 'RECAUDOS SISTECREDITO', 'RECAUDO FLAMINGO', 'TARJETA FLAMINGO', 'RECAUDO PAYU', 'RECAUDO AVON', 'RECAUDO YANBAL', 'RECAUDO SAFETYPAY', 'RECAUDO PAGO DIGITAL', 'RECAUDO AXA COLPATRIA', 'REC OPTIMA DE URABA', 'RECAUDOS JARDINES DEL EDEN', 'REC SAN GABRIEL', 'REC CAPILLAS DE LA FE', 'REC PROEXEQUIALES', 'REC FUNERARIA GOMEZ', 'RECAUDO FUNERARIA SAN JUAN BAUTISTA', 'REC FUN INMACULADA', 'REC SANTA CLARA', 'REC UBIKME', 'REC PREVER', 'F COMPASION', 'REC RESURGIR', 'REC ESCOBAR', 'REC CABLEMAS', 'FUNERARIA RENACER', 'RECAUDO LOS OLIVOS', 'REC SAN GABRIEL MED', 'REC ANORI', 'REC OFFCORSS', 'FUN SAN VICENTE', 'RECAUDO SAN NICOLAS', 'RECAUDO LOS LAURELES', 'FUNERARIA NAZARENO', 'REC FERIA DE CREDITO', 'REC RITUALES FUNERARIOS', 'RECAUDOS TELEVID', 'RECAUDO AVANTEL', 'RECAUDO MI BOLSILLO', 'RECAUDO ESIKA LBEL CYZONE', 'RECAUDO PROSEGUR', 'RECAUDO DUPREE', 'RECAUDO DATACREDITO', 'RECAUDO WOM', 'RECAUDO EMONKEY', 'RECAUDO SERVICREDITO', 'REC LEONISA', 'REC MARKETING PERSON', 'RECAUDO RUTTA', 'REC COBELEN', 'RECAUDOS PAYVALIDA', 'RECAUDO ELECTROFERIA', 'REC DIRECTV', 'RECAUDO CREDITOS PLANAUTOS', 'RECAUDO PLUSS TV', 'GEOLINK', 'REC EPAYCO', 'REC VISION GERENCIAL', 'RECAUDO ELECTROBELLO', 'SOAT', 'REC INTERACTUAR', 'RECAUDO FUNERARIA NAZARENO ', 'REC NATURA', 'RECAUDOS LOGUIN', 'RECAUDOS LA ESPERANZA', 'RECAUDO REAL HUMAN', 'RECAUDO PACIFIKA', 'RECAUDOS CARMEL', 'RECAUDO YERBABUENA', 'PAGOS COBELEN', 'RECAUDO COORDIUTIL', 'RIO APP', 'PAGOS INTERACTURA', 'PAGOS PROVEEDORES INTERACTUAL', 'RECAUDO GEOLINK', 'RECAUDO AVON (NACIONAL)', 'DEPOSITO BANCO AGRARIO', 'PAGO CARTERA BANCO AGRARIO', 'PAGO TARJETA DE CREDITO BANCO AGRARIO', 'RECAUDO BANCO AGRARIO', 'RETIRO BANCO AGRARIO', 'RECAUDO OFICINAS LOTICOLOMBIA', 'REC VENDEDORES LOTICOLOMBIA', 'REC LIMA CIA', 'REC PRODUCTOS CARIBE', 'SAN GABRIEL MED', 'REC ELECTROBELLO', 'REC LAURELES', 'LA MEDIA NARANJA', 'FINAMIGA', 'FEM AHORROS', 'FEM CONVENIOS', 'FEM FERIAS', 'FEM CREDITOS', 'PAGOS FEMFUTURO', 'PAGO GIROS', 'RETIROS BET PLAY', 'OKI')}

# Convertir el diccionario products en una cadena separada por '|'
product_str = '|'.join(products.get('OTROS'))

# Nombre de los reportes xls generados.
# filename = 'nota'
# Nombre de la carpeta donde serán almacenados los reportes generados.
folder_name = 'notas'

global today

# Ruta al reporte de notas (carpeta de OneDrive)
report_path = join(onedrive_path, report_name)

# Ruta al archivo plantilla PDF_Nota
template_path = join(templates_path, template)

# Declaración de una variable para la fecha de hoy
today = dt.datetime.today().strftime("%d-%m-%Y")

# Crear la carpeta por fecha.
folder = join(path, "reports", today)
# Si la carpeta por fecha no existe la crea.
if not exists(folder):
    mkdir(folder)

# Crear carpeta para notas
folder = join(folder, folder_name)
if not exists(folder):
    mkdir(folder)

# Lectura del reporte de la hoja de reporte de notas.
with open(report_path, mode='rb') as fp:
    df = pd.read_excel(fp, sheet_name='Hoja 1', dtype={'Nro Caso': str})

# Filtrar por columna 'Nro Caso' diferente a 'canal' o 'CANAL'
df_filtered = df[~df['Nro Caso'].str.contains('CANAL|canal', na=False)]

# Crear un Dataframe para productos diferentes a REC EPM EN LINEA o PAGO CONFAMA
df_filtered = df_filtered[df_filtered['Producto'].str.contains(product_str, na=False)]
df_filtered.reset_index(drop=True, inplace=True)

# Plantilla de Nota
# Workbook de la plantilla nota
workbook = oxl.load_workbook(template_path)
# Seleccionar la primera hoja como worksheet
worksheet = workbook.active

cells = {
    "no": ("Id. Nota", 7),
    "tipo": ("Naturaleza", 8),
    "cc": ("Responsable", 9),
    "oficina": ("Oficina", 10),
    "centro": ("Oficina", 11),
    "valor": ("Valor", 12),
    "producto": ("Producto", 13),
    "caso": ("Nro Caso", 14),
    "obs": ("Observaciones", 16)
}

# Iterar sobre df_filtered para extraer los valores y escribirlos en un archivo basado en
# la plantilla de template
for index, row in df_filtered.iterrows():
    # Insertar la fecha de la nota "fecha": (today, 5)
    worksheet.cell(5, 3, f'{today}')
    for col_name, _row in cells.values():
        if col_name == 'Id. Nota':
            filename = row[col_name]
            worksheet.cell(row=_row, column=3, value=row[col_name])
        elif col_name == 'Valor':
            worksheet.cell(row=_row, column=3, value=f'${row[col_name]:,}')
        elif col_name == 'Oficina':
            if _row == 10:
                worksheet.cell(row=_row, column=3, value=row[col_name].split('|')[1])
            if _row == 11:
                worksheet.cell(row=_row, column=3, value=row[col_name].split('|')[0])
        else:
            worksheet.cell(row=_row, column=3, value=row[col_name])
    workbook = oxl.load_workbook(template_path) #Para volver abrir el workbook y cogerla ahi mismo
    workbook.save(join(folder, f'{filename}_{today}.xlsx'))
    print(f'{filename}_{today} saved.')
workbook.close()
print("All notas generated")